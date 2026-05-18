"""
Test Case Template - Importable Python Module
=============================================
IMPORT this module instead of copying code inline.

Usage in create_*.py scripts:
    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
    from configs.tc_template import (
        save_xlsx_local, create_tc_spreadsheet, create_multi_sheet_tc_spreadsheet,
        sanitize_text, sanitize_tc, TOTAL_COLS, FONT_NAME, TIME_EST_FORMULA
    )

DO NOT copy constants or functions from this file into your script.
DO NOT redefine TIME_EST_FORMULA, TOTAL_COLS, FONT_NAME, sanitize_text, sanitize_tc.
"""

import json, os, sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Load env vars
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
from configs.env_loader import load_env, get_drive_folder_id, get_oauth_token_path, get_lark_drive_folder_id
load_env()

# ============================================================
# CONSTANTS - DO NOT REDEFINE IN YOUR SCRIPT
# ============================================================

# Total columns in the template (A-O = 15)
TOTAL_COLS = 15
# Default font for all cells (xlsx + Google Sheets)
FONT_NAME = 'Arial'
# Time Est formula for B6
# Auto-converts minutes to "X giờ Y phút" format
# WARNING: "phút" and "giờ" MUST have Vietnamese diacritics. Never write "phut" or "gio".
TIME_EST_FORMULA = '=IF(SUM(I10:I999)<60,SUM(I10:I999)&" phút / "&E5&" cases",IF(MOD(SUM(I10:I999),60)=0,INT(SUM(I10:I999)/60)&" giờ / "&E5&" cases",INT(SUM(I10:I999)/60)&" giờ "&MOD(SUM(I10:I999),60)&" phút / "&E5&" cases"))'


# ============================================================
# SANITIZE FUNCTIONS
# ============================================================

def sanitize_text(text):
    """Replace AI-generated special characters with standard ones."""
    if not isinstance(text, str):
        return text
    text = text.replace('\u2014', '-')   # em dash
    text = text.replace('\u2013', '-')   # en dash
    text = text.replace('\u2192', '=>')  # arrow
    text = text.replace('\u2018', "'")   # left single quote
    text = text.replace('\u2019', "'")   # right single quote
    text = text.replace('\u201c', '"')   # left double quote
    text = text.replace('\u201d', '"')   # right double quote
    return text


def sanitize_tc(tc):
    """Sanitize all text fields in a test case dict."""
    result = {
        'id': tc.get('id', ''),
        'desc': sanitize_text(tc.get('desc', '')),
        'precond': sanitize_text(tc.get('precond', '')),
        'steps': sanitize_text(tc.get('steps', '')),
        'expected': sanitize_text(tc.get('expected', '')),
        'time_est': tc.get('time_est', ''),
    }
    if tc.get('merge_with_prev'):
        result['merge_with_prev'] = True
    return result


# ============================================================
# GOOGLE CREDENTIALS
# ============================================================

def get_google_creds():
    """Load and refresh Google OAuth credentials from token file."""
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request

    token_file = get_oauth_token_path()
    if not os.path.exists(token_file):
        raise FileNotFoundError(
            f"OAuth token not found at {token_file}\n"
            "Run: python3 configs/setup-oauth.py"
        )
    with open(token_file, 'r') as f:
        token_data = json.load(f)
    creds = Credentials(
        token=token_data['access_token'],
        refresh_token=token_data.get('refresh_token') or os.environ.get('GOOGLE_OAUTH_REFRESH_TOKEN', ''),
        token_uri=token_data['token_uri'],
        client_id=token_data['client_id'],
        client_secret=token_data['client_secret'],
        scopes=token_data['scopes']
    )
    if creds.expired or not creds.valid:
        creds.refresh(Request())
        token_data['access_token'] = creds.token
        if creds.expiry:
            token_data['expiry_date'] = creds.expiry.isoformat()
        with open(token_file, 'w') as f:
            json.dump(token_data, f, indent=2)
    return creds


# ============================================================
# XLSX WRITING HELPERS
# ============================================================

def _write_tc_data_rows_xlsx(ws, tc_list, start_row, dv, bold_font, base_font, center_align, bold_top_align, top_align, thin_border):
    """Write TC data rows to an openpyxl worksheet, handling merge_with_prev."""
    current_row = start_row
    merge_ranges = []
    merge_start = None

    for tc in tc_list:
        tc = sanitize_tc(tc)
        is_merge = tc.get('merge_with_prev', False)

        if is_merge:
            if merge_start is None:
                merge_start = current_row - 1
        else:
            if merge_start is not None:
                merge_ranges.append((merge_start, current_row - 1))
                merge_start = None

        ws.cell(row=current_row, column=1, value=tc['id']).font = bold_font
        ws.cell(row=current_row, column=1).alignment = center_align
        if not is_merge:
            ws.cell(row=current_row, column=2, value=tc['desc']).font = bold_font
            ws.cell(row=current_row, column=2).alignment = bold_top_align
            ws.cell(row=current_row, column=3, value=tc.get('precond', '')).alignment = top_align
        ws.cell(row=current_row, column=4, value=tc.get('steps', '')).alignment = top_align
        ws.cell(row=current_row, column=5, value=tc.get('expected', '')).alignment = top_align
        status_cell = ws.cell(row=current_row, column=6, value='NOT START')
        status_cell.alignment = center_align
        dv.add(status_cell)
        time_est = tc.get('time_est', '')
        if time_est != '':
            ws.cell(row=current_row, column=9, value=time_est)
        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(row=current_row, column=c)
            cell.border = thin_border
            if not cell.font.bold:
                cell.font = base_font
        current_row += 1

    if merge_start is not None:
        merge_ranges.append((merge_start, current_row - 1))

    for m_start, m_end in merge_ranges:
        ws.merge_cells(f'B{m_start}:B{m_end}')
        ws.merge_cells(f'C{m_start}:C{m_end}')

    return current_row, merge_ranges


def _write_sheet_to_ws(ws, sheet_name, header_values, sections, data_rows, total_tc, time_est_str=None):
    """Write header + column headers + data rows into an openpyxl worksheet."""
    base_font = Font(name=FONT_NAME)
    bold_font = Font(name=FONT_NAME, bold=True)
    header_fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')
    section_fill = PatternFill(start_color='ECE2FE', end_color='ECE2FE', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_align = Alignment(vertical='top', wrap_text=True)
    bold_top_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )
    status_colors = {'Passed': 'D9F5D6', 'Failed': 'FBBFBC', 'Not start': 'E1EAFF', 'Cancel': 'DEE0E3'}

    created_date = header_values.get('created_date', datetime.now().strftime('%Y-%m-%d'))
    b6_value = time_est_str if time_est_str else TIME_EST_FORMULA
    header_data = [
        ['Link DOC:', header_values.get('link_doc', ''), '', 'Passed', f'=COUNTIF(F10:F{total_tc + 20},D1)'],
        ['Link Figma:', header_values.get('link_figma', ''), '', 'Failed', f'=COUNTIF(F10:F{total_tc + 20},D2)'],
        ['Create by', 'QA Ops Suite', '', 'Not start', f'=COUNTIF(F10:F{total_tc + 20},D3)'],
        ['Create Date', created_date, '', 'Cancel', f'=COUNTIF(F10:F{total_tc + 20},D4)'],
        ['', '', '', 'Testcases', '=SUM(E1:E4)'],
        ['Time Est (1 round):', b6_value],
    ]
    for r_idx, row_data in enumerate(header_data, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = bold_font if c_idx == 1 else base_font
        if r_idx <= 4:
            d_cell = ws.cell(row=r_idx, column=4)
            color_key = d_cell.value
            if color_key in status_colors:
                d_cell.fill = PatternFill(start_color=status_colors[color_key], end_color=status_colors[color_key], fill_type='solid')
    ws.cell(row=6, column=1).font = bold_font
    ws.cell(row=6, column=2).font = bold_font

    row7 = ['Testcase ID', 'Testcase Description', 'Pre - Condition', 'Test Procedures', '', 'Status', '', 'Bug ID', 'Time Est', 'isAuto', 'Build #1()', '', '', 'Status #1', 'Notes']
    row8 = ['', '', '', 'Steps to Perform', 'Steps Expected Result', 'Build STG', 'Build PRD', '', '', '', 'Device #1()', 'Device #2()', 'Device #3()', '', '']
    for c_idx, val in enumerate(row7, 1):
        cell = ws.cell(row=7, column=c_idx, value=val)
        cell.font = bold_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    for c_idx, val in enumerate(row8, 1):
        cell = ws.cell(row=8, column=c_idx, value=val)
        cell.font = bold_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    ws.merge_cells('D7:E7'); ws.merge_cells('F7:G7'); ws.merge_cells('K7:M7')
    for col_letter in ['A', 'B', 'C', 'H', 'I', 'J', 'N', 'O']:
        ws.merge_cells(f'{col_letter}7:{col_letter}8')

    current_row = 9
    dv = DataValidation(type='list', formula1='"PASSED,FAILED,NOT START,CANCEL"', allow_blank=True)
    dv.showDropDown = False
    ws.add_data_validation(dv)

    if isinstance(sections, str):
        ws.merge_cells(f'A{current_row}:O{current_row}')
        cell = ws.cell(row=current_row, column=1, value=sanitize_text(sections))
        cell.font = bold_font; cell.fill = section_fill; cell.alignment = center_align
        current_row += 1
        current_row, _ = _write_tc_data_rows_xlsx(
            ws, data_rows, current_row, dv,
            bold_font, base_font, center_align, bold_top_align, top_align, thin_border
        )
    else:
        for section_title, section_data in sections:
            ws.merge_cells(f'A{current_row}:O{current_row}')
            cell = ws.cell(row=current_row, column=1, value=sanitize_text(section_title))
            cell.font = bold_font; cell.fill = section_fill; cell.alignment = center_align
            current_row += 1
            current_row, _ = _write_tc_data_rows_xlsx(
                ws, section_data, current_row, dv,
                bold_font, base_font, center_align, bold_top_align, top_align, thin_border
            )

    col_widths = {'A': 12, 'B': 35, 'C': 28, 'D': 35, 'E': 35,
                  'F': 14, 'G': 14, 'H': 14, 'I': 11, 'J': 11,
                  'K': 14, 'L': 14, 'M': 14, 'N': 14, 'O': 16}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = 'A9'
    return current_row


# ============================================================
# PUBLIC API - Save xlsx locally
# ============================================================

def save_xlsx_local(filepath, sheet_name, header_values, data_rows, sections, total_tc, time_est_str=None):
    """
    Save test cases as .xlsx file locally.

    Args:
        filepath: Full path to save .xlsx file
        sheet_name: Sheet tab name
        header_values: dict with keys: link_doc, link_figma, created_date
        data_rows: list of TC dicts (used when sections is a single string)
        sections: list of tuples [(section_title, [tc_dicts]), ...] OR single string
        total_tc: Number of test cases
        time_est_str: Optional override for Time Est formula in B6
    Returns: filepath
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    _write_sheet_to_ws(ws, sheet_name, header_values, sections, data_rows, total_tc, time_est_str)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath


# ============================================================
# PUBLIC API - Create single-sheet Google Sheets
# ============================================================

def create_tc_spreadsheet(title, sheet_name, header_values, data_rows, sections, total_tc, time_est_str=None, output_dir=None):
    """
    Save .xlsx locally, then upload to Google Sheets.
    Returns: (spreadsheet_id, url, local_filepath)
    """
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'results')
    local_filepath = os.path.join(output_dir, f'{title}.xlsx')
    save_xlsx_local(local_filepath, sheet_name, header_values, data_rows, sections, total_tc, time_est_str)
    print(f"[OK] Local backup saved: {local_filepath}")

    try:
        from googleapiclient.discovery import build
        creds = get_google_creds()
        sheets_svc = build('sheets', 'v4', credentials=creds)
        drive_svc = build('drive', 'v3', credentials=creds)
        FOLDER_ID = get_drive_folder_id()

        if not FOLDER_ID:
            print("[INFO] GOOGLE_DRIVE_FOLDER_ID not set. Creating new Drive folder...")
            folder_metadata = {'name': 'QA Ops Suite - Test Cases', 'mimeType': 'application/vnd.google-apps.folder'}
            folder = drive_svc.files().create(body=folder_metadata, fields='id').execute()
            FOLDER_ID = folder.get('id')
            print(f"[OK] Created Drive folder: https://drive.google.com/drive/folders/{FOLDER_ID}")

        sp = sheets_svc.spreadsheets().create(body={'properties': {'title': title, 'defaultFormat': {'textFormat': {'fontFamily': FONT_NAME}}}}).execute()
        sp_id = sp['spreadsheetId']
        f = drive_svc.files().get(fileId=sp_id, fields='parents').execute()
        drive_svc.files().update(fileId=sp_id, addParents=FOLDER_ID, removeParents=','.join(f.get('parents', [])), fields='id').execute()
        sheets_svc.spreadsheets().batchUpdate(spreadsheetId=sp_id, body={'requests': [
            {'updateSheetProperties': {'properties': {'sheetId': 0, 'title': sheet_name}, 'fields': 'title'}}
        ]}).execute()

        values, section_rows, merge_b_c_ranges = _build_sheet_values(sheet_name, header_values, sections, data_rows, total_tc, time_est_str)
        last_row = len(values)
        sheets_svc.spreadsheets().values().update(
            spreadsheetId=sp_id, range=f'{sheet_name}!A1:O{last_row}',
            valueInputOption='USER_ENTERED', body={'values': values}
        ).execute()

        requests = _build_format_requests(0, last_row, section_rows, merge_b_c_ranges)
        sheets_svc.spreadsheets().batchUpdate(spreadsheetId=sp_id, body={'requests': requests}).execute()

        url = f"https://docs.google.com/spreadsheets/d/{sp_id}"
        print(f"[OK] Google Sheets created: {url}")
        return sp_id, url, local_filepath
    except Exception as e:
        print(f"[WARN] Google Sheets upload failed: {e}")
        print(f"[OK] Local .xlsx file is still available at: {local_filepath}")
        return None, None, local_filepath


# ============================================================
# PUBLIC API - Create multi-sheet Google Sheets
# ============================================================

def create_multi_sheet_tc_spreadsheet(title, sheets_data, output_dir=None):
    """
    Create a formatted test case spreadsheet with MULTIPLE SHEETS.
    Each sheet has its own header (rows 1-6) with COUNTIF formulas.
    TC_ID resets per sheet.

    Args:
        title: Spreadsheet title
        sheets_data: list of dicts with keys: sheet_name, header_values, data_rows, sections, total_tc, time_est_str
        output_dir: Directory to save local .xlsx file
    Returns: (spreadsheet_id, url, local_filepath)
    """
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'results')
    local_filepath = os.path.join(output_dir, f'{title}.xlsx')
    os.makedirs(os.path.dirname(local_filepath), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    for sd in sheets_data:
        ws = wb.create_sheet(title=sd['sheet_name'])
        _write_sheet_to_ws(
            ws, sd['sheet_name'], sd.get('header_values', {}),
            sd['sections'], sd.get('data_rows', []),
            sd['total_tc'], sd.get('time_est_str')
        )
    wb.save(local_filepath)
    print(f"[OK] Local backup saved: {local_filepath}")

    try:
        from googleapiclient.discovery import build
        creds = get_google_creds()
        sheets_svc = build('sheets', 'v4', credentials=creds)
        drive_svc = build('drive', 'v3', credentials=creds)
        FOLDER_ID = get_drive_folder_id()

        if not FOLDER_ID:
            folder_metadata = {'name': 'QA Ops Suite - Test Cases', 'mimeType': 'application/vnd.google-apps.folder'}
            folder = drive_svc.files().create(body=folder_metadata, fields='id').execute()
            FOLDER_ID = folder.get('id')

        sheet_props = [{'properties': {'title': sd['sheet_name']}} for sd in sheets_data]
        sp = sheets_svc.spreadsheets().create(body={
            'properties': {'title': title, 'defaultFormat': {'textFormat': {'fontFamily': FONT_NAME}}},
            'sheets': sheet_props
        }).execute()
        sp_id = sp['spreadsheetId']
        f = drive_svc.files().get(fileId=sp_id, fields='parents').execute()
        drive_svc.files().update(fileId=sp_id, addParents=FOLDER_ID, removeParents=','.join(f.get('parents', [])), fields='id').execute()

        sheet_id_map = {s['properties']['title']: s['properties']['sheetId'] for s in sp['sheets']}
        all_requests = []

        for sd in sheets_data:
            sname = sd['sheet_name']
            sid = sheet_id_map[sname]
            values, section_rows, merge_b_c_ranges = _build_sheet_values(
                sname, sd.get('header_values', {}), sd['sections'],
                sd.get('data_rows', []), sd['total_tc'], sd.get('time_est_str')
            )
            last_row = len(values)
            sheets_svc.spreadsheets().values().update(
                spreadsheetId=sp_id, range=f"'{sname}'!A1:O{last_row}",
                valueInputOption='USER_ENTERED', body={'values': values}
            ).execute()
            all_requests.extend(_build_format_requests(sid, last_row, section_rows, merge_b_c_ranges))

        if all_requests:
            sheets_svc.spreadsheets().batchUpdate(spreadsheetId=sp_id, body={'requests': all_requests}).execute()

        url = f"https://docs.google.com/spreadsheets/d/{sp_id}"
        print(f"[OK] Google Sheets created ({len(sheets_data)} sheets): {url}")
        return sp_id, url, local_filepath
    except Exception as e:
        print(f"[WARN] Google Sheets upload failed: {e}")
        print(f"[OK] Local .xlsx file is still available at: {local_filepath}")
        return None, None, local_filepath


# ============================================================
# INTERNAL - Build Google Sheets values array
# ============================================================

def _build_sheet_values(sheet_name, header_values, sections, data_rows, total_tc, time_est_str=None):
    """Build the values array for Google Sheets API. Returns (values, section_rows, merge_b_c_ranges)."""
    created_date = header_values.get('created_date', datetime.now().strftime('%Y-%m-%d'))
    b6_value = time_est_str if time_est_str else TIME_EST_FORMULA
    values = [
        ['Link DOC:', header_values.get('link_doc', ''), '', 'Passed', '=COUNTIF(F10:F;D1)'] + ['']*(TOTAL_COLS - 5),
        ['Link Figma:', header_values.get('link_figma', ''), '', 'Failed', '=COUNTIF(F10:F;D2)'] + ['']*(TOTAL_COLS - 5),
        ['Create by', 'QA Ops Suite', '', 'Not start', '=COUNTIF(F10:F;D3)'] + ['']*(TOTAL_COLS - 5),
        ['Create Date', created_date, '', 'Cancel', '=COUNTIF(F10:F;D4)'] + ['']*(TOTAL_COLS - 5),
        ['', '', '', 'Testcases', '=SUM(E1:E4)'] + ['']*(TOTAL_COLS - 5),
        ['Time Est (1 round):', b6_value] + ['']*(TOTAL_COLS - 2),
        ['Testcase ID', 'Testcase Description', 'Pre - Condition', 'Test Procedures', '', 'Status', '', 'Bug ID', 'Time Est', 'isAuto', 'Build #1()', '', '', 'Status #1', 'Notes'],
        ['', '', '', 'Steps to Perform', 'Steps Expected Result', 'Build STG', 'Build PRD', '', '', '', 'Device #1()', 'Device #2()', 'Device #3()', '', ''],
    ]
    section_rows = []
    merge_b_c_ranges = []

    def _append_tc_rows(tc_list):
        merge_start = None
        for tc in tc_list:
            tc = sanitize_tc(tc)
            is_merge = tc.get('merge_with_prev', False)
            if is_merge:
                if merge_start is None:
                    merge_start = len(values) - 1
            else:
                if merge_start is not None:
                    merge_b_c_ranges.append((merge_start, len(values) - 1))
                    merge_start = None
            row = [''] * TOTAL_COLS
            row[0] = tc['id']
            if not is_merge:
                row[1] = tc['desc']
                row[2] = tc.get('precond', '')
            row[3] = tc.get('steps', '')
            row[4] = tc.get('expected', '')
            row[5] = 'NOT START'
            row[8] = tc.get('time_est', '')
            values.append(row)
        if merge_start is not None:
            merge_b_c_ranges.append((merge_start, len(values) - 1))

    if isinstance(sections, str):
        values.append([sanitize_text(sections)] + ['']*(TOTAL_COLS - 1))
        section_rows.append(len(values) - 1)
        _append_tc_rows(data_rows)
    else:
        for section_title, section_data in sections:
            values.append([sanitize_text(section_title)] + ['']*(TOTAL_COLS - 1))
            section_rows.append(len(values) - 1)
            _append_tc_rows(section_data)

    return values, section_rows, merge_b_c_ranges


# ============================================================
# INTERNAL - Build Google Sheets formatting requests
# ============================================================

def _build_format_requests(sid, last_row, section_rows, merge_b_c_ranges):
    """Build formatting requests for Google Sheets API."""
    requests = []
    # Header rows 7-8
    requests.append({'repeatCell': {'range': {'sheetId': sid, 'startRowIndex': 6, 'endRowIndex': 8, 'startColumnIndex': 0, 'endColumnIndex': TOTAL_COLS}, 'cell': {'userEnteredFormat': {'backgroundColor': {'red': 0.8, 'green': 1.0, 'blue': 1.0}, 'textFormat': {'bold': True, 'fontFamily': FONT_NAME}, 'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE'}}, 'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)'}})
    # Section rows
    for sr in section_rows:
        requests.append({'repeatCell': {'range': {'sheetId': sid, 'startRowIndex': sr, 'endRowIndex': sr+1, 'startColumnIndex': 0, 'endColumnIndex': TOTAL_COLS}, 'cell': {'userEnteredFormat': {'backgroundColor': {'red': 0.925, 'green': 0.886, 'blue': 0.996}, 'textFormat': {'bold': True, 'fontFamily': FONT_NAME}, 'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE'}}, 'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)'}})
        requests.append({'mergeCells': {'range': {'sheetId': sid, 'startRowIndex': sr, 'endRowIndex': sr+1, 'startColumnIndex': 0, 'endColumnIndex': TOTAL_COLS}, 'mergeType': 'MERGE_ALL'}})
    # Multi-result TC merges
    for m_start, m_end in merge_b_c_ranges:
        requests.append({'mergeCells': {'range': {'sheetId': sid, 'startRowIndex': m_start, 'endRowIndex': m_end + 1, 'startColumnIndex': 1, 'endColumnIndex': 2}, 'mergeType': 'MERGE_ALL'}})
        requests.append({'mergeCells': {'range': {'sheetId': sid, 'startRowIndex': m_start, 'endRowIndex': m_end + 1, 'startColumnIndex': 2, 'endColumnIndex': 3}, 'mergeType': 'MERGE_ALL'}})
    # Col A bold+center, Col B bold+wrap, Col C-O wrap, Status center
    requests.append({'repeatCell': {'range': {'sheetId': sid, 'startRowIndex': 8, 'endRowIndex': last_row, 'startColumnIndex': 0, 'endColumnIndex': 1}, 'cell': {'userEnteredFormat': {'textFormat': {'bold': True, 'fontFamily': FONT_NAME}, 'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE'}}, 'fields': 'userEnteredFormat(textFormat,horizontalAlignment,verticalAlignment)'}})
    requests.append({'repeatCell': {'range': {'sheetId': sid, 'startRowIndex': 8, 'endRowIndex': last_row, 'startColumnIndex': 1, 'endColumnIndex': 2}, 'cell': {'userEnteredFormat': {'textFormat': {'bold': True, 'fontFamily': FONT_NAME}, 'wrapStrategy': 'WRAP', 'verticalAlignment': 'TOP'}}, 'fields': 'userEnteredFormat(textFormat,wrapStrategy,verticalAlignment)'}})
    requests.append({'repeatCell': {'range': {'sheetId': sid, 'startRowIndex': 8, 'endRowIndex': last_row, 'startColumnIndex': 2, 'endColumnIndex': TOTAL_COLS}, 'cell': {'userEnteredFormat': {'wrapStrategy': 'WRAP', 'verticalAlignment': 'TOP'}}, 'fields': 'userEnteredFormat(wrapStrategy,verticalAlignment)'}})
    requests.append({'repeatCell': {'range': {'sheetId': sid, 'startRowIndex': 8, 'endRowIndex': last_row, 'startColumnIndex': 5, 'endColumnIndex': 7}, 'cell': {'userEnteredFormat': {'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE', 'wrapStrategy': 'WRAP'}}, 'fields': 'userEnteredFormat(horizontalAlignment,verticalAlignment,wrapStrategy)'}})
    # Header area bold
    requests.append({'repeatCell': {'range': {'sheetId': sid, 'startRowIndex': 0, 'endRowIndex': 6, 'startColumnIndex': 0, 'endColumnIndex': 1}, 'cell': {'userEnteredFormat': {'textFormat': {'bold': True, 'fontFamily': FONT_NAME}}}, 'fields': 'userEnteredFormat(textFormat)'}})
    requests.append({'repeatCell': {'range': {'sheetId': sid, 'startRowIndex': 5, 'endRowIndex': 6, 'startColumnIndex': 0, 'endColumnIndex': 2}, 'cell': {'userEnteredFormat': {'textFormat': {'bold': True, 'fontFamily': FONT_NAME}}}, 'fields': 'userEnteredFormat(textFormat)'}})
    # Status summary colors
    for row_idx, color in [(0, {'red':0.851,'green':0.961,'blue':0.839}), (1, {'red':0.984,'green':0.749,'blue':0.737}), (2, {'red':0.882,'green':0.918,'blue':1.0}), (3, {'red':0.871,'green':0.878,'blue':0.890})]:
        requests.append({'repeatCell': {'range': {'sheetId': sid, 'startRowIndex': row_idx, 'endRowIndex': row_idx+1, 'startColumnIndex': 3, 'endColumnIndex': 4}, 'cell': {'userEnteredFormat': {'backgroundColor': color, 'horizontalAlignment': 'RIGHT'}}, 'fields': 'userEnteredFormat(backgroundColor,horizontalAlignment)'}})
    # Column widths
    for col_idx, px in [(0,80),(1,250),(2,200),(3,250),(4,250),(5,100),(6,100),(7,100),(8,80)]:
        requests.append({'updateDimensionProperties': {'range': {'sheetId': sid, 'dimension': 'COLUMNS', 'startIndex': col_idx, 'endIndex': col_idx+1}, 'properties': {'pixelSize': px}, 'fields': 'pixelSize'}})
    # Freeze + borders
    requests.append({'updateSheetProperties': {'properties': {'sheetId': sid, 'gridProperties': {'frozenRowCount': 8}}, 'fields': 'gridProperties.frozenRowCount'}})
    requests.append({'updateBorders': {'range': {'sheetId': sid, 'startRowIndex': 6, 'endRowIndex': last_row, 'startColumnIndex': 0, 'endColumnIndex': TOTAL_COLS}, 'top': {'style':'SOLID','color':{'red':0.7,'green':0.7,'blue':0.7}}, 'bottom': {'style':'SOLID','color':{'red':0.7,'green':0.7,'blue':0.7}}, 'left': {'style':'SOLID','color':{'red':0.7,'green':0.7,'blue':0.7}}, 'right': {'style':'SOLID','color':{'red':0.7,'green':0.7,'blue':0.7}}, 'innerHorizontal': {'style':'SOLID','color':{'red':0.7,'green':0.7,'blue':0.7}}, 'innerVertical': {'style':'SOLID','color':{'red':0.7,'green':0.7,'blue':0.7}}}})
    # Header merges
    requests.append({'mergeCells': {'range': {'sheetId': sid, 'startRowIndex': 6, 'endRowIndex': 7, 'startColumnIndex': 3, 'endColumnIndex': 5}, 'mergeType': 'MERGE_ALL'}})
    requests.append({'mergeCells': {'range': {'sheetId': sid, 'startRowIndex': 6, 'endRowIndex': 7, 'startColumnIndex': 5, 'endColumnIndex': 7}, 'mergeType': 'MERGE_ALL'}})
    requests.append({'mergeCells': {'range': {'sheetId': sid, 'startRowIndex': 6, 'endRowIndex': 7, 'startColumnIndex': 10, 'endColumnIndex': 13}, 'mergeType': 'MERGE_ALL'}})
    for col in [0, 1, 2, 7, 8, 9, 13, 14]:
        requests.append({'mergeCells': {'range': {'sheetId': sid, 'startRowIndex': 6, 'endRowIndex': 8, 'startColumnIndex': col, 'endColumnIndex': col+1}, 'mergeType': 'MERGE_ALL'}})
    # Data validation
    data_start = 8
    for col_start, col_end in [(5, 6), (6, 7)]:
        requests.append({'setDataValidation': {'range': {'sheetId': sid, 'startRowIndex': data_start, 'endRowIndex': last_row, 'startColumnIndex': col_start, 'endColumnIndex': col_end}, 'rule': {'condition': {'type': 'ONE_OF_LIST', 'values': [{'userEnteredValue': 'PASSED'}, {'userEnteredValue': 'FAILED'}, {'userEnteredValue': 'NOT START'}, {'userEnteredValue': 'CANCEL'}]}, 'showCustomUi': True, 'strict': True}}})
    return requests
