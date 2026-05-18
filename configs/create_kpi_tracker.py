"""Tạo file Team KPI Tracker xlsx với 5 sheets (4 cá nhân + 1 summary).

Framework: 5 KPIs chuẩn cho cả 4 người, target cá nhân hóa theo level.
Chi tiết: .claude/docs/team-kpi.md

Usage:
    python3 configs/create_kpi_tracker.py [--month YYYY-MM] [--out PATH]
"""
from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# 5 KPIs chuẩn cả team (theo Section 2 trong team-kpi.md)
# ---------------------------------------------------------------------------

# Mỗi KPI: (idx, name, weight%_default, direction, formula, source)
# weight%_default = trọng số áp dụng cho Mid/Junior. Lead có weight scale ×0.8 + KPI #6 (20%).
KPIS_BASE = [
    (1, "Bug Leak Rate (cá nhân)", 30, "down",
     "% bugs lọt PRD = (Bugs found in PRD / Bugs found in STG) cho features mình handle",
     "Lark Bug Board (filter Owner=bạn, Environment=Production vs Staging)"),
    (2, "Automation Testcase", 30, "up",
     "Số auto test (Appium/Newman/Selenium) viết mới + chạy stable trong tháng",
     "Git commits + CI/CD test reports"),
    (3, "New Feature SBH/tháng", 20, "up",
     "Số tính năng mới hoàn thành test trong tháng (mọi product)",
     "Lark Task Board (Type=Feature, Status=Done, QA_Sign_Off in tháng)"),
    (4, "TC mới cho Features mới (Auto + Manual)", 10, "up",
     "Tổng TC viết mới (auto + manual) cho features mới trong tháng",
     "Google Sheets TC (Created_At in tháng, Feature in new_features_list)"),
    (5, "API Testing", 10, "up",
     "Số API test (Postman/Newman) viết mới + execute trong tháng",
     "Postman workspace + Newman run reports + Git commits"),
]

# KPI #6 — Lead only. Khi person có role "lead" => append KPI này, scale 5 KPI trên ×0.8.
KPI_LEAD_EXTRA = (
    6, "Quản lý & Mentor (Lead)", 20, "up",
    ("Composite score 3 chỉ số: "
     "(a) 1:1 với team ≥ 8 buổi/tháng (4 người × 2 buổi); "
     "(b) Code/TC review trong SLA 24h ≥ 95%; "
     "(c) Mentor activity ≥ 4/tháng (sprint review chair, retro chair, training session, pair-program)"),
    "Calendar logs + Git PR review timestamps + Team meeting minutes",
)
LEAD_SCALE = 0.8  # 5 KPIs Lead × 0.8 = 80%, cộng KPI #6 (20%) = 100%

# Target + baseline + note theo từng người (key = KPI idx)
# is_lead=True => weights scaled ×0.8 + KPI #6 added
PERSON_TARGETS = {
    "Tường": {
        "role": "Lead — Senior, Mentor, Selenium Framework Owner",
        "is_lead": True,
        "kpis": {
            1: {"baseline": "12%", "target": "< 5% (target excellent)", "note": "Per-person leak rate. Tier: 0%=5/5, <5%=4/5, <10%=3/5, <15%=2/5, <20%=1/5, ≥20%=0/5"},
            2: {"baseline": "3", "target": "≥ 5", "note": "Selenium scripts viết mới + review code Mid/Junior"},
            3: {"baseline": "2", "target": "≥ 2", "note": "Oversight: ensure team test xong ≥ 2 features/tháng"},
            4: {"baseline": "25", "target": "≥ 30", "note": "Review TC team + viết TC critical phần Lead phụ trách"},
            5: {"baseline": "2", "target": "≥ 3", "note": "Review collections của Dung + spot-check Postman"},
            6: {"baseline": "(T5 baseline)", "target": "≥ 90%", "note": "KPI riêng cho Lead — đảm bảo Tường có thời gian quản lý team, không bị overload bởi execution KPI"},
        },
    },
    "Ngọc Anh": {
        "role": "Mid — Mobile Auto Specialist target (Appium + Java)",
        "is_lead": False,
        "kpis": {
            1: {"baseline": "18%", "target": "< 5% (target excellent)", "note": "Per-person leak rate. Tier: 0%=5/5, <5%=4/5, <10%=3/5, <15%=2/5, <20%=1/5, ≥20%=0/5"},
            2: {"baseline": "1", "target": "≥ 2", "note": "Appium scripts mới, chạy stable. Q2 milestone: 5+ scripts cuối Q"},
            3: {"baseline": "2", "target": "≥ 2", "note": "Features được giao test xong + sign-off"},
            4: {"baseline": "28", "target": "≥ 30", "note": "Auto + manual TC cho features mới (target Q2 AI-gen >60%)"},
            5: {"baseline": "0", "target": "≥ 1 collection", "note": "Cross-skill Postman cơ bản (theo roadmap Q2)"},
        },
    },
    "Dung": {
        "role": "Mid — API Auto Specialist target (Postman/Newman + Java)",
        "is_lead": False,
        "kpis": {
            1: {"baseline": "10%", "target": "< 5% (target excellent)", "note": "Per-person leak rate. Tier: 0%=5/5, <5%=4/5, <10%=3/5, <15%=2/5, <20%=1/5, ≥20%=0/5"},
            2: {"baseline": "2", "target": "≥ 3", "note": "Newman tests mới, chạy stable. Q2 milestone: 2+ products coverage"},
            3: {"baseline": "2", "target": "≥ 2", "note": "Features được giao test xong + sign-off"},
            4: {"baseline": "26", "target": "≥ 30", "note": "Auto + manual TC cho features mới (target Q2 AI-gen >60%)"},
            5: {"baseline": "2", "target": "≥ 3", "note": "Newman tests mới (chuyên môn chính)"},
        },
    },
    "Phước": {
        "role": "Junior — Hybrid QC target (Manual + AI + Java basics)",
        "is_lead": False,
        "kpis": {
            1: {"baseline": "20%", "target": "< 5% (target excellent)", "note": "Per-person leak rate. Tier: 0%=5/5, <5%=4/5, <10%=3/5, <15%=2/5, <20%=1/5, ≥20%=0/5"},
            2: {"baseline": "0", "target": "≥ 1 (pair)", "note": "Pair test + đọc hiểu auto code (Q2 chưa tự viết)"},
            3: {"baseline": "2", "target": "≥ 2", "note": "Features được giao test xong + sign-off"},
            4: {"baseline": "30", "target": "≥ 30", "note": "Manual chính + AI-assist (target Q2 AI usage >80%)"},
            5: {"baseline": "3", "target": "≥ 5 endpoints", "note": "Postman manual: tự gửi request + đọc response"},
        },
    },
}


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

FONT_NAME = "Arial"
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
COL_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
SUMMARY_HEADER_FILL = PatternFill("solid", fgColor="C00000")

WHITE_BOLD = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
BOLD = Font(name=FONT_NAME, size=11, bold=True)
NORMAL = Font(name=FONT_NAME, size=11)
SMALL = Font(name=FONT_NAME, size=10, italic=True, color="595959")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def build_person_sheet(ws: Worksheet, name: str, month_label: str) -> None:
    """Build sheet KPI cá nhân.

    Mid/Junior: 5 KPIs (30/30/20/10/10) tổng 100%.
    Lead: 6 KPIs — 5 KPIs scale ×0.8 (24/24/16/8/8) + KPI #6 Quản lý 20% = 100%.
    """
    person = PERSON_TARGETS[name]
    role = person["role"]
    is_lead = person.get("is_lead", False)

    # Build effective KPI list for this person
    if is_lead:
        scaled_kpis = [
            (idx, kname, weight * LEAD_SCALE, direction, formula, source)
            for (idx, kname, weight, direction, formula, source) in KPIS_BASE
        ]
        kpis_for_person = scaled_kpis + [KPI_LEAD_EXTRA]
    else:
        kpis_for_person = list(KPIS_BASE)

    # Row 1 - title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    cell = ws.cell(row=1, column=1, value=f"KPI Tháng {month_label} — {name}")
    cell.font = WHITE_BOLD
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Row 2 - role
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    cell = ws.cell(row=2, column=1, value=role)
    cell.font = Font(name=FONT_NAME, size=11, italic=True, color="FFFFFF")
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    ws.row_dimensions[2].height = 22

    # Row 4 - column headers
    headers = [
        "#", "KPI", "Trọng số", "Hướng",
        "Baseline (T-1)", "Target (T)", "Actual (T)",
        "% đạt", "Δ vs T-1", "Note / Source",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = BOLD
        c.fill = COL_HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[4].height = 30

    # Data rows
    start_row = 5
    weight_cells: list[str] = []
    pct_cells: list[str] = []

    for i, (idx, kname, weight, direction, formula, source) in enumerate(kpis_for_person):
        r = start_row + i
        cfg = person["kpis"][idx]

        full_note = f"Formula: {formula}\nSource: {source}\nNote: {cfg['note']}"

        ws.cell(row=r, column=1, value=idx).alignment = CENTER
        ws.cell(row=r, column=2, value=kname).alignment = LEFT
        wcell = ws.cell(row=r, column=3, value=weight / 100.0)
        wcell.number_format = "0%"
        wcell.alignment = CENTER
        ws.cell(row=r, column=4, value="↑ Càng cao càng tốt" if direction == "up" else "↓ Càng thấp càng tốt").alignment = CENTER
        ws.cell(row=r, column=5, value=cfg["baseline"]).alignment = CENTER
        ws.cell(row=r, column=6, value=cfg["target"]).alignment = CENTER
        ws.cell(row=r, column=7, value="").alignment = CENTER  # Actual
        ws.cell(row=r, column=8, value="").alignment = CENTER  # % đạt
        ws.cell(row=r, column=9, value="").alignment = CENTER  # Δ
        ws.cell(row=r, column=10, value=full_note).alignment = LEFT

        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            if col_idx != 3:
                cell.font = NORMAL
            cell.border = BORDER

        weight_cells.append(f"C{r}")
        pct_cells.append(f"H{r}")

    # Weighted Score row (after all KPIs — count varies for Lead vs Mid/Junior)
    total_row = start_row + len(kpis_for_person)
    ws.cell(row=total_row, column=1, value="").border = BORDER
    cell = ws.cell(row=total_row, column=2, value="WEIGHTED SCORE")
    cell.font = BOLD
    cell.alignment = LEFT
    cell.fill = TOTAL_FILL
    cell.border = BORDER

    weight_sum_formula = f"=SUM({weight_cells[0]}:{weight_cells[-1]})"
    cell = ws.cell(row=total_row, column=3, value=weight_sum_formula)
    cell.number_format = "0%"
    cell.font = BOLD
    cell.alignment = CENTER
    cell.fill = TOTAL_FILL
    cell.border = BORDER

    for col in range(4, 8):
        c = ws.cell(row=total_row, column=col, value="")
        c.fill = TOTAL_FILL
        c.border = BORDER

    score_formula = (
        f"=IFERROR(SUMPRODUCT({pct_cells[0]}:{pct_cells[-1]},"
        f"{weight_cells[0]}:{weight_cells[-1]}), \"\")"
    )
    cell = ws.cell(row=total_row, column=8, value=score_formula)
    cell.number_format = "0%"
    cell.font = BOLD
    cell.alignment = CENTER
    cell.fill = TOTAL_FILL
    cell.border = BORDER

    for col in [9, 10]:
        c = ws.cell(row=total_row, column=col, value="")
        c.fill = TOTAL_FILL
        c.border = BORDER

    # Note phân loại
    note_row = total_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    cell = ws.cell(
        row=note_row,
        column=1,
        value=(
            "Phân loại: ≥110% Vượt xuất sắc · 100-110% Vượt KPI · 90-100% Đạt · "
            "80-90% Cận đạt (review) · <80% Không đạt (1:1 + action plan). "
            "Cách tính % đạt: KPI ↑ = Actual/Target, KPI ↓ = Target/Actual, cap 120%."
        ),
    )
    cell.font = SMALL
    cell.alignment = LEFT

    # Column widths
    widths = {
        "A": 5, "B": 38, "C": 10, "D": 22,
        "E": 16, "F": 16, "G": 14, "H": 10, "I": 12, "J": 50,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "C5"


def build_summary_sheet(ws: Worksheet, month_label: str) -> None:
    """Team Summary: Bug Ticket Rate aggregate + Weighted Score per person."""
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    cell = ws.cell(row=1, column=1, value=f"TEAM SUMMARY — Tháng {month_label}")
    cell.font = WHITE_BOLD
    cell.fill = SUMMARY_HEADER_FILL
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Section 1: Bug Leak Rate per person (Target < 5%)
    ws.cell(row=3, column=1, value="1. BUG LEAK RATE PER PERSON (Target < 5%)").font = BOLD
    headers1 = ["Người", "Bugs in PRD", "Bugs in STG", "Leak Rate", "Target", "Tier (điểm/5)"]
    for i, h in enumerate(headers1, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = BOLD
        c.fill = COL_HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    def tier_formula(rate_cell: str) -> str:
        return (
            f'=IF({rate_cell}="","",'
            f'IF({rate_cell}=0,"5/5 (100%)",'
            f'IF({rate_cell}<0.05,"4/5 (80%)",'
            f'IF({rate_cell}<0.10,"3/5 (60%)",'
            f'IF({rate_cell}<0.15,"2/5 (40%)",'
            f'IF({rate_cell}<0.20,"1/5 (20%)","0/5 (0%) - FAIL"))))))'
        )

    persons = ["Tường", "Ngọc Anh", "Dung", "Phước"]
    for i, p in enumerate(persons):
        r = 5 + i
        ws.cell(row=r, column=1, value=p).alignment = LEFT
        ws.cell(row=r, column=2, value="").alignment = CENTER
        ws.cell(row=r, column=3, value="").alignment = CENTER
        ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/C{r},\"\")").number_format = "0.0%"
        ws.cell(row=r, column=4).alignment = CENTER
        ws.cell(row=r, column=5, value="< 5%").alignment = CENTER
        ws.cell(row=r, column=6, value=tier_formula(f"D{r}")).alignment = CENTER

        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
            if col != 4:
                ws.cell(row=r, column=col).font = NORMAL

    # Team aggregate row
    team_row = 5 + len(persons)
    ws.cell(row=team_row, column=1, value="TEAM AGGREGATE").font = BOLD
    ws.cell(row=team_row, column=1).fill = TOTAL_FILL
    ws.cell(row=team_row, column=2, value=f"=SUM(B5:B{team_row-1})").alignment = CENTER
    ws.cell(row=team_row, column=3, value=f"=SUM(C5:C{team_row-1})").alignment = CENTER
    ws.cell(row=team_row, column=4, value=f"=IFERROR(B{team_row}/C{team_row},\"\")").number_format = "0.0%"
    ws.cell(row=team_row, column=4).alignment = CENTER
    ws.cell(row=team_row, column=4).font = BOLD
    ws.cell(row=team_row, column=5, value="< 5%").alignment = CENTER
    ws.cell(row=team_row, column=6, value=tier_formula(f"D{team_row}")).alignment = CENTER
    ws.cell(row=team_row, column=6).font = BOLD
    for col in range(1, 7):
        ws.cell(row=team_row, column=col).border = BORDER
        if col != 1:
            ws.cell(row=team_row, column=col).fill = TOTAL_FILL

    # Section 2: Weighted Score per person
    sec2_row = team_row + 3
    ws.cell(row=sec2_row, column=1, value="2. WEIGHTED SCORE PER PERSON").font = BOLD

    headers2 = ["Người", "Role", "Weighted Score", "Phân loại", "Note"]
    for i, h in enumerate(headers2, start=1):
        c = ws.cell(row=sec2_row + 1, column=i, value=h)
        c.font = BOLD
        c.fill = COL_HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # Score row = 5 (start_row) + số KPI. Lead có 6 KPIs => H11; Mid/Junior 5 KPIs => H10.
    roles = {
        "Tường": "Lead — Senior",
        "Ngọc Anh": "Mid — Mobile Auto",
        "Dung": "Mid — API Auto",
        "Phước": "Junior — Hybrid QC",
    }

    for i, p in enumerate(persons):
        r = sec2_row + 2 + i
        is_lead = PERSON_TARGETS[p].get("is_lead", False)
        score_row = 5 + (len(KPIS_BASE) + 1 if is_lead else len(KPIS_BASE))
        ws.cell(row=r, column=1, value=p).alignment = LEFT
        ws.cell(row=r, column=2, value=roles[p]).alignment = LEFT
        ws.cell(row=r, column=3, value=f"='{p}'!H{score_row}").number_format = "0%"
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=3).font = BOLD
        score_cell = f"C{r}"
        rank_formula = (
            f'=IF({score_cell}="","",'
            f'IF({score_cell}>=1.1,"Vượt xuất sắc",'
            f'IF({score_cell}>=1.0,"Vượt KPI",'
            f'IF({score_cell}>=0.9,"Đạt",'
            f'IF({score_cell}>=0.8,"Cận đạt","Không đạt")))))'
        )
        ws.cell(row=r, column=4, value=rank_formula).alignment = CENTER
        ws.cell(row=r, column=5, value="").alignment = LEFT
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER

    # Note hướng dẫn
    note_row = sec2_row + 2 + len(persons) + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    cell = ws.cell(
        row=note_row,
        column=1,
        value=(
            "Section 1 - Bug Leak Rate per person: Cuối tháng fill 'Bugs in PRD' (bug lọt) + 'Bugs in STG' (bug catch trước release). "
            "Tier tự cập nhật: <5%=4/5, <10%=3/5, <15%=2/5, <20%=1/5, ≥20%=0/5, =0% là 5/5. "
            "Section 2 - Weighted Score: tổng từ sheet cá nhân, đã includes Bug Leak Rate. "
            "Hướng dẫn chi tiết từng KPI: xem sheet 'Hướng dẫn đạt KPI'."
        ),
    )
    cell.font = SMALL
    cell.alignment = LEFT

    widths = {"A": 16, "B": 22, "C": 18, "D": 18, "E": 50, "F": 14}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ---------------------------------------------------------------------------
# Guidance sheet — "Hướng dẫn đạt KPI"
# ---------------------------------------------------------------------------

PROCESS_RULES = [
    ("1. Regression Test Estimation",
     "Trước khi bắt đầu test ROUND 1 cho 1 feature, bắt buộc phải est được thời gian regression test. "
     "Regression có thể nằm ở round 4 hoặc 5 nhưng phải est trước. Em ghi est vào plan-tc + report cho Lead."),
    ("2. Chuẩn release (Release Standard)",
     "Feature được phép release khi: 0 bug Critical/High open + ≤ 2 bug Medium open + ≤ 10 bug Low open. "
     "Nếu vượt mức → KHÔNG release, đẩy lùi sprint hoặc xin extension."),
    ("3. Quyền quyết định release",
     "Mỗi QC tự quyết định release/không release cho feature mình handle, không cần chờ Lead approve. "
     "Trách nhiệm đi kèm: nếu release ra mà bugs lọt PRD → ảnh hưởng KPI Bug Leak Rate cá nhân."),
    ("4. QC Focus = Chất lượng SẢN PHẨM",
     "QC giờ chỉ focus chất lượng sản phẩm (không phải chất lượng team). Phong cách: solution-oriented — "
     "không chỉ report bug, mà đề xuất giải pháp nâng cao chất lượng (process, tooling, automation, AI...). "
     "Defensive QC bị coi là hết thời."),
]

KPI_GUIDANCE = [
    {
        "title": "KPI #1 — Bug Leak Rate (cá nhân)  [30% Mid/Junior, 24% Lead]",
        "muc_tieu": "Mục tiêu: < 5% bugs lọt từ STG ra PRD. Đo per-person theo feature handle.",
        "can_lam": (
            "Cần làm gì:\n"
            "  • Test coverage tốt: phủ đủ 4 nhóm (UI display, data type, data display, interaction/validation)\n"
            "  • Negative + edge case + boundary cho mọi feature\n"
            "  • Regression test ở round 4-5 (đã est từ round 1)\n"
            "  • Risk-based testing: ưu tiên high-impact area (payment, login, data integrity)\n"
            "  • Pair test với dev cho phần phức tạp\n"
            "  • Dùng AI/automation hỗ trợ rà soát coverage gap\n"
            "  • Sign-off feature theo Chuẩn release (mục I.2)"
        ),
        "do_cach_nao": (
            "Đo cách nào:\n"
            "  • Bugs in PRD = bug Source=Customer/Production của feature mình QA, trong tháng\n"
            "  • Bugs in STG = bug catch ở STG/UAT của feature mình QA, trong tháng\n"
            "  • Leak Rate = Bugs PRD / Bugs STG\n"
            "  • Source: Lark Bug Board, filter Owner + Environment"
        ),
        "extra": (
            "Scoring tier:\n"
            "  • 0%      = 5/5 (100%) — stretch goal\n"
            "  • < 5%    = 4/5 (80%)  — excellent (target)\n"
            "  • 5-10%   = 3/5 (60%)  — good\n"
            "  • 10-15%  = 2/5 (40%)  — warning\n"
            "  • 15-20%  = 1/5 (20%)  — poor\n"
            "  • ≥ 20%   = 0/5 (0%)   — fail (1:1 + action plan)"
        ),
    },
    {
        "title": "KPI #2 — Automation Testcase  [30% Mid/Junior, 24% Lead]",
        "muc_tieu": "Mục tiêu: viết auto test mới, chạy stable trong CI. Tường ≥5, NA ≥2 Appium, Dung ≥3 Newman, Phước ≥1 pair.",
        "can_lam": (
            "Cần làm gì:\n"
            "  • Identify TC nào nên auto: high-frequency execution + stable spec\n"
            "  • Viết theo framework đã có (Selenium/Appium/Newman)\n"
            "  • Chạy CI ≥ 3 lần liên tiếp pass = 'stable'\n"
            "  • Phước: pair với senior, đọc hiểu code có sẵn trước\n"
            "  • Q2: AI-assisted test creation > 60%"
        ),
        "do_cach_nao": (
            "Đo cách nào:\n"
            "  • Count auto test files/methods MERGED vào main trong tháng\n"
            "  • 'Stable' = pass run đầu trong CI, không retry, không flaky\n"
            "  • Source: Git commits + CI/CD test reports"
        ),
        "extra": (
            "Pitfalls (KHÔNG tính):\n"
            "  • Test flaky (retry mới pass)\n"
            "  • Auto-port từ legacy không refactor\n"
            "  • Test fail CI nhưng pass local"
        ),
    },
    {
        "title": "KPI #3 — New Feature SBH/tháng  [20% Mid/Junior, 16% Lead]",
        "muc_tieu": "Mục tiêu: ≥ 2 feature/tháng được test xong + sign-off (≈ 1 feature/sprint).",
        "can_lam": (
            "Cần làm gì:\n"
            "  • Nhận feature từ sprint planning, plan TC trong 1-2 ngày đầu sprint\n"
            "  • Test execution: round 1 → 2 (re-test) → 3+ (regression)\n"
            "  • Ghi log đầy đủ: bugs, retest, blocker\n"
            "  • Sign-off feature theo Chuẩn release\n"
            "  • Lead (Tường): oversight, không phải tự test 2 feature"
        ),
        "do_cach_nao": (
            "Đo cách nào:\n"
            "  • Lark Task Board: Type=Feature + Status=Done + QA_Sign_Off in tháng + Owner=bạn\n"
            "  • 'Done' = đã release hoặc đã sign-off cho UAT/PRD"
        ),
        "extra": "Lưu ý: feature nhỏ (<5 TC) gộp 2 = 1 feature. Feature lớn (>50 TC) tách 2 phase = 2 feature.",
    },
    {
        "title": "KPI #4 — TC mới cho Features mới  [10% Mid/Junior, 8% Lead]",
        "muc_tieu": "Mục tiêu: ≥ 30 TC/tháng (auto + manual) cho feature mới.",
        "can_lam": (
            "Cần làm gì:\n"
            "  • Dùng QA Ops Suite /cook generate TC từ specs/Figma — Q2 AI-gen >60%\n"
            "  • Theo 4-part framework\n"
            "  • Negative + edge case bắt buộc cho mọi positive TC\n"
            "  • Tường: review TC team + viết TC critical phụ trách"
        ),
        "do_cach_nao": (
            "Đo cách nào:\n"
            "  • Google Sheets TC: Created_At in tháng + Feature in new_features_list + Owner=bạn\n"
            "  • Count TC rows (không count section header rows)"
        ),
        "extra": "Pitfalls: TC duplicate, TC color check, TC chỉ 'app không crash' — KHÔNG tính.",
    },
    {
        "title": "KPI #5 — API Testing  [10% Mid/Junior, 8% Lead]",
        "muc_tieu": "Mục tiêu: API coverage. Dung ≥3 Newman, NA ≥1 collection cross-skill, Phước ≥5 endpoints, Tường ≥3 review.",
        "can_lam": (
            "Cần làm gì:\n"
            "  • Postman collection: 1 collection = 1 module (Auth, Order, Payment...)\n"
            "  • Newman test: chạy được trong CI (Dung ưu tiên)\n"
            "  • Coverage: positive + negative (4xx, 5xx) + edge (timeout, big payload)\n"
            "  • Phước: manual gửi request, đọc response, log issue"
        ),
        "do_cach_nao": (
            "Đo cách nào:\n"
            "  • Count Postman collection mới + Newman runs trong tháng\n"
            "  • Source: Postman workspace + Newman reports + Git commits"
        ),
        "extra": "Phước count theo endpoint manual (≥5), không cần collection. NA cross-skill ≥1 collection (roadmap Q2).",
    },
    {
        "title": "KPI #6 — Quản lý & Mentor (Lead only)  [20% Lead]",
        "muc_tieu": "Mục tiêu: Composite score ≥ 90% — đảm bảo Lead có thời gian quản lý team.",
        "can_lam": (
            "Cần làm gì (Tường):\n"
            "  • 1:1 với từng team member ≥ 8 buổi/tháng (4 người × 2 buổi)\n"
            "  • Code/TC review trong SLA 24h ≥ 95%\n"
            "  • Mentor activity ≥ 4/tháng: chair sprint review, chair retro, tổ chức training, pair-program với junior\n"
            "  • Đề xuất ≥ 1 process improvement / quarter"
        ),
        "do_cach_nao": (
            "Đo cách nào:\n"
            "  • Calendar logs (1:1 sessions)\n"
            "  • Git PR review timestamps (SLA)\n"
            "  • Team meeting minutes (mentor activity)"
        ),
        "extra": "Composite score = trung bình 3 chỉ số (a, b, c). Target ≥ 90%.",
    },
]


def build_guidance_sheet(ws: Worksheet) -> None:
    """Build sheet 'Hướng dẫn đạt KPI' — guide cho team về cách đạt từng KPI."""
    rules_title_fill = PatternFill("solid", fgColor="D9E1F2")
    rules_title_font = Font(name=FONT_NAME, size=11, bold=True, color="1F4E78")
    section_fill = PatternFill("solid", fgColor="C00000")
    section_font = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
    kpi_title_fill = PatternFill("solid", fgColor="2E75B6")
    kpi_title_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=FONT_NAME, size=10)
    body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    r = 1
    # Title
    cell = ws.cell(row=r, column=1, value="HƯỚNG DẪN ĐẠT KPI — Team QA SBH")
    cell.font = WHITE_BOLD
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    ws.row_dimensions[r].height = 36
    r += 2

    # Section I: Process Rules
    cell = ws.cell(row=r, column=1, value="I. QUY TRÌNH BẮT BUỘC (mọi cấp)")
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = body_align
    ws.row_dimensions[r].height = 28
    r += 2

    for title, body in PROCESS_RULES:
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = rules_title_font
        cell.fill = rules_title_fill
        cell.alignment = body_align
        r += 1
        cell = ws.cell(row=r, column=1, value=body)
        cell.font = body_font
        cell.alignment = body_align
        ws.row_dimensions[r].height = 60
        r += 2

    r += 1
    # Section II: KPI detail
    cell = ws.cell(row=r, column=1, value="II. CHI TIẾT TỪNG KPI — LÀM GÌ ĐỂ ĐẠT")
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = body_align
    ws.row_dimensions[r].height = 28
    r += 2

    for kpi in KPI_GUIDANCE:
        cell = ws.cell(row=r, column=1, value=kpi["title"])
        cell.font = kpi_title_font
        cell.fill = kpi_title_fill
        cell.alignment = body_align
        ws.row_dimensions[r].height = 24
        r += 1
        for key in ["muc_tieu", "can_lam", "do_cach_nao", "extra"]:
            cell = ws.cell(row=r, column=1, value=kpi[key])
            cell.font = body_font
            cell.alignment = body_align
            # Auto-height handled by openpyxl when wrap_text + multi-line
            ws.row_dimensions[r].height = max(20, kpi[key].count("\n") * 16 + 20)
            r += 1
        r += 1

    # Footer
    cell = ws.cell(row=r + 1, column=1, value="Tham chiếu chi tiết: .claude/docs/team-kpi.md (framework đầy đủ)")
    cell.font = SMALL
    cell.alignment = body_align
    cell = ws.cell(row=r + 2, column=1, value="Created by: QA Ops Suite — Team Lead: Tường")
    cell.font = SMALL
    cell.alignment = body_align

    ws.column_dimensions["A"].width = 110


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def build_workbook(month_label: str, output_path: Path) -> Path:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Summary first
    ws_sum = wb.create_sheet(title="Team Summary")
    build_summary_sheet(ws_sum, month_label)

    # Guidance sheet (right after summary)
    ws_guide = wb.create_sheet(title="Hướng dẫn đạt KPI")
    build_guidance_sheet(ws_guide)

    # Person sheets
    for name in ["Tường", "Ngọc Anh", "Dung", "Phước"]:
        ws = wb.create_sheet(title=name)
        build_person_sheet(ws, name, month_label)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--month", default=None, help="Month label (e.g. T5/2026). Default: current.")
    parser.add_argument("--out", default=None, help="Output xlsx path.")
    args = parser.parse_args()

    if args.month:
        month_label = args.month
    else:
        now = datetime.now()
        month_label = f"T{now.month}/{now.year}"

    if args.out:
        out = Path(args.out)
    else:
        repo_root = Path(__file__).resolve().parents[1]
        out = repo_root / "results" / "team-kpi" / "team-kpi-tracker-2026.xlsx"

    result = build_workbook(month_label, out)
    print(f"Created: {result}")
    print(f"Month: {month_label}")
    print(f"Sheets: Team Summary | Hướng dẫn đạt KPI | Tường | Ngọc Anh | Dung | Phước")


if __name__ == "__main__":
    main()
